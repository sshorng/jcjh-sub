# -*- coding: utf-8 -*-
"""監考表匯出範例：每人一張 xlsx（假資料）"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(__file__), "監考表匯出-範例_每人一張.xlsx")

TITLE = "臺北市立建成國民中學114學年度第一學期第一次段考監考表（範例）"
DAY1 = "10月13日(星期一)"
DAY2 = "10月14日(星期二)"
PERIODS = [
    ("1", "08:30", "09:15"),
    ("2", "09:25", "10:10"),
    ("3", "10:20", "11:05"),
    ("4", "11:15", "12:00"),
    ("5", "13:20", "14:05"),
    ("6", "14:15", "15:00"),
    ("7", "15:15", "16:00"),
    ("1", "08:30", "09:15"),
    ("2", "09:25", "10:10"),
    ("3", "10:20", "11:05"),
    ("4", "11:15", "12:00"),
]
GRADE = {
    7: ["自習", "社會", "自習", "數學", "自習", "國文", "自習", "自習", "英語", "自習", "生物"],
    8: ["自習", "社會", "自習", "數學", "自習", "國文", "自習", "自習", "英語", "自習", "理化"],
    9: ["自習", "社會", "自習", "數學", "自習", "國文", "自習", "自習", "英語", "自習", "自然"],
}

TEACHERS = [
    {
        "name": "王小明",
        "quota_before": 4,
        "quota_used_exam": 2,
        "quota_remain": 2,
        "slots": [
            {"text": "701", "changed": False},
            {"text": "巡堂", "changed": True},
            {"text": "", "changed": False},
            {"text": "702", "changed": False},
            {"text": "", "changed": False},
            {"text": "巡七", "changed": True},
            {"text": "", "changed": False},
            {"text": "自習", "changed": False},
            {"text": "703", "changed": True},
            {"text": "", "changed": False},
            {"text": "巡堂", "changed": True},
        ],
    },
    {
        "name": "陳美玲",
        "quota_before": 0,
        "quota_used_exam": 0,
        "quota_remain": 0,
        "slots": [
            {"text": "特殊考場", "changed": False},
            {"text": "", "changed": False},
            {"text": "801", "changed": False},
            {"text": "巡八", "changed": True},
            {"text": "", "changed": False},
            {"text": "公假", "changed": False},
            {"text": "", "changed": False},
            {"text": "", "changed": False},
            {"text": "802", "changed": False},
            {"text": "", "changed": False},
            {"text": "", "changed": False},
        ],
    },
    {
        "name": "林大文",
        "quota_before": 3,
        "quota_used_exam": 1,
        "quota_remain": 2,
        "slots": [
            {"text": "", "changed": False},
            {"text": "901", "changed": False},
            {"text": "巡九", "changed": True},
            {"text": "", "changed": False},
            {"text": "902", "changed": False},
            {"text": "", "changed": False},
            {"text": "903代", "changed": True},
            {"text": "巡堂", "changed": True},
            {"text": "", "changed": False},
            {"text": "904", "changed": False},
            {"text": "", "changed": False},
        ],
    },
]

thin = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
font_title = Font(name="標楷體", size=18, bold=True)
font_h = Font(name="標楷體", size=12)
font_cell = Font(name="標楷體", size=14)
font_chg = Font(name="標楷體", size=14, bold=True, underline="single")
font_name = Font(name="標楷體", size=16, bold=True, color="C00000")
font_note = Font(name="標楷體", size=11)
font_tip = Font(name="標楷體", size=12, bold=True)
fill_name = PatternFill("solid", fgColor="FFF2CC")
fill_header = PatternFill("solid", fgColor="F1F5F9")
fill_chg = PatternFill("solid", fgColor="FEF3C7")


def write_teacher_sheet(wb, t):
    ws = wb.create_sheet(title=t["name"][:28])
    ws.column_dimensions["A"].width = 10
    for i in range(2, 13):
        ws.column_dimensions[get_column_letter(i)].width = 9
    ws.column_dimensions["N"].width = 18
    ws.column_dimensions["O"].width = 12

    ws.merge_cells("A1:L1")
    ws["A1"] = TITLE
    ws["A1"].font = font_title
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 36

    ws.merge_cells("N1:O1")
    ws["N1"] = "姓名：" + t["name"]
    ws["N1"].font = font_name
    ws["N1"].fill = fill_name
    ws["N1"].alignment = center
    ws["N1"].border = thin
    ws["O1"].border = thin

    ws["A2"] = "日  期"
    ws.merge_cells("B2:H2")
    ws["B2"] = DAY1
    ws.merge_cells("I2:L2")
    ws["I2"] = DAY2
    for c in range(1, 13):
        cell = ws.cell(2, c)
        cell.font = font_h
        cell.alignment = center
        cell.border = thin
        cell.fill = fill_header
    ws.row_dimensions[2].height = 24

    ws["A3"] = "節次"
    for i, p in enumerate(PERIODS):
        cell = ws.cell(3, 2 + i, p[0])
        cell.font = font_h
        cell.alignment = center
        cell.border = thin
        cell.fill = fill_header
    ws["A3"].font = font_h
    ws["A3"].alignment = center
    ws["A3"].border = thin
    ws["A3"].fill = fill_header

    for gi, g in enumerate([7, 8, 9]):
        r = 4 + gi
        ws.cell(r, 1, str(g) + "年級").font = font_h
        ws.cell(r, 1).alignment = center
        ws.cell(r, 1).border = thin
        for i, subj in enumerate(GRADE[g]):
            cell = ws.cell(r, 2 + i, subj)
            cell.font = Font(name="標楷體", size=11)
            cell.alignment = center
            cell.border = thin

    ws["A7"] = "時間"
    ws["A8"] = "教師"
    for r in (7, 8):
        ws.cell(r, 1).font = font_h
        ws.cell(r, 1).alignment = center
        ws.cell(r, 1).border = thin
        ws.cell(r, 1).fill = fill_header
    for i, p in enumerate(PERIODS):
        for r, val in ((7, p[1]), (8, p[2])):
            cell = ws.cell(r, 2 + i, val)
            cell.font = Font(name="標楷體", size=10)
            cell.alignment = center
            cell.border = thin
            cell.fill = fill_header

    ws["A9"] = t["name"]
    ws["A9"].font = font_name
    ws["A9"].fill = fill_name
    ws["A9"].alignment = center
    ws["A9"].border = thin
    for i, slot in enumerate(t["slots"]):
        text = slot["text"] or None
        cell = ws.cell(9, 2 + i, text)
        cell.alignment = center
        cell.border = thin
        if slot.get("changed") and slot.get("text"):
            cell.font = font_chg
            cell.fill = fill_chg
        else:
            cell.font = font_cell
    ws.row_dimensions[9].height = 28

    ws.merge_cells("A10:L10")
    ws["A10"] = "★圖例：粗體底線（淺黃）＝異動（含空堂排班／調代／代巡）；空白＝該節無監考安排。"
    ws["A10"].font = font_tip
    ws["A10"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[10].height = 28

    qb = t["quota_before"]
    qu = t["quota_used_exam"]
    qr = t["quota_remain"]
    note = (
        "備註：\n"
        "(1)請監考老師留意各科考試時間及節次，提早至教務處領取試卷，至少提前3分鐘入班，"
        "指導學生安靜入座，鐘響前發放完畢答案卡，鐘聲響起立即發下考卷，讓學生有完整時間作答。"
        "(一致性原則詳見段考範圍表所列注意事項)\n"
        "(2)各科考試時間皆為一節課(45分鐘)，測驗開始及結束時間皆以學校鐘聲為準。\n"
        "(3)監考表原則採隨班監考方式安排呈現，部分節次因命題巡堂與考科當節迴避而適度調整，"
        "請留意監考節次與班級。\n"
        "(4)請老師專心監考，不要讓學生有作弊的機會，並請於非考試節次依本表準時入班監看自習。\n"
        "(5)校外教學期間，教師未執行課務(安排已劃底線註記)，依比例安排於本學年段考監考、"
        "監看自習或代理遺留課程【未執行的課務共 {0} 節，本次段考已安排 {1} 節，尚有 {2} 節，"
        "未執行節數將會累計於本學年度】\n"
        "　　（系統對照：未執行的課務＝安排前折抵額度；本次段考已安排＝段考期間扣額度節數；"
        "尚有＝目前剩餘額度）"
    ).format(qb, qu, qr)
    ws.merge_cells("A11:L11")
    ws["A11"] = note
    ws["A11"].font = font_note
    ws["A11"].alignment = left_top
    ws["A11"].border = thin
    ws.row_dimensions[11].height = 168

    ws["N3"] = "額度摘要"
    ws["N3"].font = font_h
    ws["N3"].fill = fill_header
    ws["N3"].border = thin
    ws["O3"] = t["name"]
    ws["O3"].border = thin
    for r, lab, val in (
        (4, "未執行的課務（安排前額度）", qb),
        (5, "本次段考已安排（期間扣額度）", qu),
        (6, "尚有（目前剩餘額度）", qr),
    ):
        ws.cell(r, 14, lab).font = Font(name="標楷體", size=10)
        ws.cell(r, 14).alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(r, 14).border = thin
        ws.cell(r, 15, val).font = Font(name="標楷體", size=14, bold=True)
        ws.cell(r, 15).alignment = center
        ws.cell(r, 15).border = thin

    ws.print_area = "A1:L11"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1


def main():
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "說明"
    ws0["A1"] = "監考表匯出範例（每人一張）"
    ws0["A1"].font = Font(name="標楷體", size=16, bold=True)
    lines = [
        "1. 每一教師一個工作表（列印＝一人一張）。",
        "2. 右上標「姓名：○○○」（黃底紅字），方便分發。",
        "3. 異動格：粗體＋底線＋淺黃底（空堂排班巡堂、調代、代字等）。",
        "4. 備註(5)：未執行課務＝安排前額度；本次已安排＝段考期間扣額度；尚有＝剩餘額度。",
        "5. 本檔假資料示範版型；正式版接課表＋空堂排班／額度帳本。",
        "6. 工作表：王小明、陳美玲、林大文。",
    ]
    for i, line in enumerate(lines):
        ws0.cell(3 + i, 1, line)
    ws0.column_dimensions["A"].width = 92

    for t in TEACHERS:
        write_teacher_sheet(wb, t)

    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
