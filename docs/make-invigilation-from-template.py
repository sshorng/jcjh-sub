# -*- coding: utf-8 -*-
"""
以桌面「114學年度第一學期第一次段考監考表.xlsx」為模板（版型不改）：
- 每人一張＝完整全校監考表複本
- 左上角（頁首）標「分發：○○○」（黑白列印可辨）
- 全校所有異動格：粗體＋底線（不加色底）
- 備註(5) 改寫並填入該員額度三欄
- 正式版：人員／課表讀系統；特殊考場、請假不自動標（空白，手動改）
"""
import os
import re
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

TEMPLATE = r"C:\Users\sshor\OneDrive\桌面\114學年度第一學期第一次段考監考表.xlsx"
OUT = os.path.join(
    os.path.dirname(__file__),
    "監考表匯出-範例_全校表每人一份.xlsx",
)

# 示範異動：全校共用（每張複本都套用）
# 正式版改由系統比對「基礎版／異動」自動標記
ALL_CHANGED_CELLS = [
    "B9", "C9", "D9",           # 戴言儒
    "G10", "I10", "L10",        # 王慈惠
    "C11",                      # 鄭惠文 巡堂
    "C12", "F12", "J12", "K12", # 黃怡君
    "B13", "C13",               # 黃書庭
    "C14",                      # 鍾筱萍
    "E16", "H16",               # 林大鈞
    "Q9", "R9", "U9", "X9",     # 黃美蘭
    "N10", "Q10", "R10", "S10", "U10", "V10",  # 張雅筑
    "O17",                      # 齊悅翔 巡八
]

DEMO = [
    {"name": "戴言儒", "quota_before": 3, "quota_used_exam": 2, "quota_remain": 1},
    {"name": "王慈惠", "quota_before": 1, "quota_used_exam": 1, "quota_remain": 0},
    {"name": "黃美蘭", "quota_before": 4, "quota_used_exam": 1, "quota_remain": 3},
]

NOTE5_OLD = re.compile(
    r"【.*?未執行的.*?節，本次段考已安排.*?節，尚有.*?節，未執行節數將會累計於本學年度】",
    re.DOTALL,
)

# 列印區外：頁首左上（不佔 A1 標題合併區）
# 用工作表 header + 可選 Y1 輔助格
HEADER_FONT = Font(name="標楷體", size=14, bold=True, underline="single")


def copy_sheet(wb, src_name, new_name):
    src = wb[src_name]
    dst = wb.copy_worksheet(src)
    safe = new_name[:31]
    base, n = safe, 1
    while safe in [s.title for s in wb.worksheets if s is not dst]:
        n += 1
        safe = (base[:28] + "_" + str(n))[:31]
    dst.title = safe
    return dst


def mark_changed(cell):
    """異動：粗體＋底線；黑白列印，不加底色"""
    if cell.value is None or str(cell.value).strip() == "":
        return
    f = copy(cell.font) if cell.font else Font(name="標楷體", size=15)
    cell.font = Font(
        name=f.name or "標楷體",
        size=f.size or 15,
        bold=True,
        underline="single",
        color=f.color,
        italic=f.italic,
        strike=f.strike,
    )


def mark_all_changes(ws, addrs):
    for addr in addrs or []:
        mark_changed(ws[addr])


def set_top_left_name(ws, name):
    """
    黑白分發識別：頁首左上「分發：姓名」
    不改 A1:X1 標題合併內容；print_area 仍 A1:X47
    """
    # oddHeader 左區塊（列印時出現在頁面左上）
    label = "分發：" + name
    try:
        ws.oddHeader.left.text = label
        ws.oddHeader.left.font = "標楷體"
        ws.oddHeader.left.size = 14
        ws.oddHeader.left.bold = True
        ws.evenHeader.left.text = label
        ws.evenHeader.left.font = "標楷體"
        ws.evenHeader.left.size = 14
        ws.evenHeader.left.bold = True
    except Exception:
        pass
    # 畫面預覽：把標題列左側可見提示寫進 A1 前方不可行（合併）
    # 改在列印區上方插入會改版型 → 禁止
    # 另：Y1（列印區外）放「分發：姓名」，螢幕可對、列印靠 header
    ws["Y1"] = label
    ws["Y1"].font = HEADER_FONT
    ws["Y1"].alignment = Alignment(horizontal="left", vertical="center")
    # 列印標題仍維持原 A1 校名（不改）


def patch_note(ws, before, used, remain):
    cell = ws["A47"]
    text = cell.value
    if text is None:
        return
    text = str(text)
    new_bracket = (
        "【未執行的課務共{0}節，本次段考已安排{1}節，尚有{2}節，"
        "未執行節數將會累計於本學年度】"
    ).format(before, used, remain)
    if NOTE5_OLD.search(text):
        text = NOTE5_OLD.sub(new_bracket, text)
    else:
        old_snip = (
            "【114.10.15~10.17畢業旅行，未執行的九年級課務共_____節，"
            "本次段考已安排_____節，尚有____節，未執行節數將會累計於本學年度】"
        )
        if old_snip in text:
            text = text.replace(old_snip, new_bracket)
        else:
            text = re.sub(r"【[^】]*未執行[^】]*】", new_bracket, text, count=1)
    cell.value = text


def personalize(ws, demo, print_area="A1:X47", all_changed=None):
    set_top_left_name(ws, demo["name"])
    mark_all_changes(ws, all_changed if all_changed is not None else ALL_CHANGED_CELLS)
    patch_note(
        ws,
        demo["quota_before"],
        demo["quota_used_exam"],
        demo["quota_remain"],
    )
    if print_area:
        ws.print_area = print_area
    # 頁面邊距略留 header 空間（不改表格本體）
    try:
        ws.page_margins.top = max(float(ws.page_margins.top or 0.5), 0.7)
    except Exception:
        pass


def main():
    if not os.path.isfile(TEMPLATE):
        raise SystemExit("找不到模板：" + TEMPLATE)

    wb = load_workbook(TEMPLATE)
    src_name = wb.sheetnames[0]

    if "說明" in wb.sheetnames:
        del wb["說明"]
    ws0 = wb.create_sheet("說明", 0)
    ws0["A1"] = "監考表匯出範例（全校表 × 每人一份）"
    ws0["A1"].font = Font(name="標楷體", size=16, bold=True)
    lines = [
        "規則：版型完全沿用模板，不改欄位／合併／列印範圍 A1:X47。",
        "每人印出＝整份全校監考表（內容相同）。",
        "分發識別（黑白）：頁首左上「分發：姓名」；Y1 同文案（螢幕對稿用，不在列印區）。",
        "異動：全校所有異動格皆粗體＋底線（不加底色，利黑白列印）。",
        "資料：人員／課表讀系統；特殊考場、請假不自動標（格留白，請手動改）。",
        "備註(5)：未執行的課務＝安排前額度；本次已安排＝段考期間扣額度；尚有＝剩餘額度。",
        "示範對象：" + "、".join(d["name"] for d in DEMO),
        "工作表「" + src_name + "」＝原始模板。",
    ]
    for i, line in enumerate(lines):
        ws0.cell(3 + i, 1, line)
    ws0.column_dimensions["A"].width = 100

    src = wb[src_name]
    print_area = src.print_area or "A1:X47"
    if "!" in str(print_area):
        print_area = str(print_area).split("!", 1)[-1].replace("$", "")

    # 原表也先套「全校異動」樣式，方便對照
    mark_all_changes(src, ALL_CHANGED_CELLS)

    for d in DEMO:
        sheet = copy_sheet(wb, src_name, d["name"])
        personalize(sheet, d, print_area=print_area, all_changed=ALL_CHANGED_CELLS)

    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
